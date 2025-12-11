
import random
import tkinter as tk
import openpyxl    #Si no está instalado, usar pip install openpyxl
import matplotlib.pyplot as grafico  #Si no está instalado, usar pip install matplotlib
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

def clean(window):  #Limpia la ventana de todos sus widgets
    for widget in window.winfo_children():
        widget.pack_forget()

def save_statistics(mode, name, attempts, won):  #Funcion para guardar las estadisticas en un archivo excel
    excel_file = openpyxl.load_workbook("statistics.xlsx")
    Hoja1 = excel_file["Solo_mode"]
    Hoja2 = excel_file["Two_players_mode"]
    if mode == "Solo":
        Hoja1.append([name,attempts, won])
    else:
        Hoja2.append([name,attempts, won])
    excel_file.save("statistics.xlsx")

def save_info(window,mode, attempts, won): #Pantalla para guardar las estadisticas, esta es llamada desde las pantallas de victoria y derrota
    clean(window)
    
    frame_game = tk.Frame(window,bg="lightblue")
    frame_game.pack(fill="both", expand=True)
    entry = tk.Entry(frame_game, font=("Arial", 18))
    entry.pack(pady=20)
    entry.focus()
    label = tk.Label(frame_game, text="Enter your name to save your statistics", font=("Arial", 24), bg="lightblue")
    label.pack(pady=20)
    def save():
        player_name = entry.get()
        save_statistics(mode, player_name, attempts, won) #Llama a la funcion save_statistics con el nombre del jugador
        back_to_menu(window)
    
    btn_save = tk.Button(window, text="Save", font=("Arial", 12), 
                      command=lambda:save()) #Ese boton llama a la funcion save
    btn_save.pack(pady=5)

def wining_message(window, mode,attempts): #Pantalla de victoria
    clean(window)
    
    gif = tk.PhotoImage(file="victory.gif")
    label = tk.Label(window, image=gif,text = "Congratulations! You guessed the number!", compound="top", font=("Arial", 24))
    label.image = gif 
    label.pack()

    boton_menu = tk.Button(window, text="Save stat", font=("Arial", 18), command=lambda: save_info(window,mode,attempts,"win"))
    boton_menu.pack(pady=50)

def losing_message(window, mode, attempts): #Pantalla de derrota
    clean(window)
    
    im = tk.PhotoImage(file="lose.gif")
    label = tk.Label(window, image=im,text = "Sorry, you lost the game.", compound="top", font=("Arial", 24))
    label.image = im
    label.pack()

    boton_menu = tk.Button(window, text="Save stat", font=("Arial", 18), command=lambda: save_info(window,mode,attempts,"lose"))
    boton_menu.pack(pady=5)



def game(window,n_to_guess,max_attempts,mode): #Funcion principal del juego
    attempts = 0 
    wining_status = False
    inf, sup = 1, 1000 #Limites iniciales para la pista de mayor/menor, se actualizan segun los intentos

    clean(window)

    frame_game = tk.Frame(window,bg="lightblue")
    frame_game.pack(fill="both", expand=True)

    label = tk.Label(frame_game, text="Introduce your number", font=("Arial", 24), bg="lightblue")
    label.pack(pady=20)

    
    entry = tk.Entry(frame_game, font=("Arial", 18))
    entry.pack(pady=20)
    entry.focus()

    def proceed(): #Funcion que se llama al presionar el boton "Send", en la que se procesa el intento
        nonlocal attempts, wining_status, inf, sup,mode
        if attempts < max_attempts and not wining_status: #Esto en verdad sobra, ya que en el caso de haber ganado
                                                          # o haber llegado al maximo de intentos, se vuelve al menu
            attempts += 1
            guess = int(entry.get())

            if guess < n_to_guess:
                inf = max(inf,guess + 1) 
                text = f"Intento {attempts} de {max_attempts}. \n Adivina el numero entre {inf} y {sup}: "    
                label.config(text=text + "El numero es mayor")
            elif guess > n_to_guess:
                sup = min(sup,guess - 1)
                text = f"Intento {attempts} de {max_attempts}. \n Adivina el numero entre {inf} y {sup}: "      
                label.config(text=text + "El numero es menor")
            else:
                wining_status = True
                wining_message(window,mode,attempts)
            
            if attempts == max_attempts and not wining_status:
                losing_message(window,mode,attempts)
        
    
    btn_send = tk.Button(window, text="Send", font=("Arial", 12), 
                      command=proceed) #Ese boton llama a la funcion proceed
    btn_send.pack(pady=5)
            
        
def back_to_menu(window): #Funcion para volver al menu principal
        clean(window)

        screen = tk.Frame(window)
        message = tk.Label(screen, text="Welcome to the Guess the Number Game!", font=("Arial", 24))
        message.pack(pady=20)
        screen.pack(fill ="both", expand = True)

        botons = tk.Frame(window)
        boton1 = tk.Button(botons, text="1.Solo mode", font=("Arial", 18),command = lambda:solo_mode(window))
        boton1.grid(column=0, row=0, padx=20, pady=20)

        boton2 = tk.Button(botons, text="2.Two players mode", font=("Arial", 18), command=lambda:two_players_mode(window))
        boton2.grid(column=0, row=1, padx=20, pady=20)

        boton3 = tk.Button(botons, text="3.Check statistics", font=("Arial", 18))
        boton3.grid(column=0, row=2, padx=20, pady=20)

        boton4 = tk.Button(botons, text="4.Exit", font=("Arial", 18), command=window.quit)
        boton4.grid(column=0, row=3, padx=20, pady=20)
        botons.pack()  
        screen.pack()

def botons_difficulty(window,n_to_guess,mode): #Pantalla para seleccionar la dificultad, es llamada desde los modos de juego
    #El clean de la ventana se hace en las funciones que llaman a esta
    frame_game = tk.Frame(window,bg="lightblue")
    frame_game.pack(fill="both", expand=True)

    label = tk.Label(frame_game, text="Modo solitario", font=("Arial", 24), bg="lightblue")
    label.pack(pady=20)

    botons_difficulty = tk.Frame(frame_game, bg="lightblue")
    botons_easy = tk.Button(botons_difficulty, text="1.Easy (20 attempts)", command = lambda: game(window,n_to_guess,20,mode), font=("Arial", 18))
    botons_easy.grid(column=0, row=0, padx=20, pady=20)

    botons_medium = tk.Button(botons_difficulty, text="2.Medium (12 attempts)", command = lambda: game(window,n_to_guess,12,mode), font=("Arial", 18))
    botons_medium.grid(column=0, row=1, padx=20, pady=20)

    botons_hard = tk.Button(botons_difficulty, text="3.Hard (5 attempts)", command = lambda: game(window,n_to_guess,5,mode), font=("Arial", 18))
    botons_hard.grid(column=0, row=2, padx=20, pady=20)

    botons_difficulty.pack()



def solo_mode(window): #Funcion para el modo solitario
    n_to_guess = random.randint(1, 1000)

    clean(window)
 
    botons_difficulty(window,n_to_guess,"Solo")


def two_players_mode(window): #Funcion para el modo de dos jugadores

    clean(window)
 
    frame_game = tk.Frame(window,bg="lightblue")
    frame_game.pack(fill="both", expand=True)

    label = tk.Label(frame_game, text="Introduzca el numero a adivinar", font=("Arial", 24), bg="lightblue")
    label.pack(pady=20)

    entry = tk.Entry(frame_game, font=("Arial", 18))
    entry.pack(pady=20)
    entry.focus()

    def get_number(): #Funcion que se llama al presionar el boton "Send", en la que se obtiene el numero a adivinar
        n_to_guess = int(entry.get())
        clean(window)
    
        botons_difficulty(window,n_to_guess,"Two_players") #Llama a la pantalla de seleccion de dificultad

    
    btn_send = tk.Button(window, text="Send", font=("Arial", 12), 
                      command=lambda: get_number())
    btn_send.pack(pady=5)

def count(data): #Funcion para contar las estadisticas a partir de los datos leidos del excel
    wins = 0
    loses = 0
    attempts_list = []
    for row in data:
        if row[2] == "win":
            wins += 1
        else:
            loses += 1
        attempts_list.append(row[1])
    attempts_mean = sum(attempts_list) / len(attempts_list)
    return wins, loses, attempts_mean


def check_statistics(): #Funcion para mostrar las estadisticas en una nueva ventana
    excel_file = openpyxl.load_workbook("statistics.xlsx")
    Hoja1 = excel_file["Solo_mode"]
    Hoja2 = excel_file["Two_players_mode"]
    def leer_hoja(hoja):
        datos = []
        for fila in hoja.iter_rows(values_only=True):
            datos.append(fila)
        return datos

    datos1 = leer_hoja(Hoja1)
    datos2 = leer_hoja(Hoja2)

    win1, lose1, attempts_mean1 = count(datos1[1:])
    win2, lose2, attempts_mean2 = count(datos2[1:])

    window = tk.Tk()
    window.title("Game Statistics")
    window.geometry("1000x700")

    screen = tk.Frame(window)
    screen.pack(fill ="both", expand = True)

    graphics_frame = tk.Frame(screen, bd=2, relief="groove")
    graphics_frame.pack(side=tk.TOP, fill="both", expand=True, pady=5)

    x1,x2 = ["Wins", "Loses"], ["Wins", "Loses"]
    y1,y2 = [win1, lose1], [win2, lose2]

    fig = Figure(figsize=(10, 5), dpi=100)
    ax1, ax2, ax3, ax4 = fig.add_subplot(221), fig.add_subplot(222), fig.add_subplot(223), fig.add_subplot(224)

    ax1.bar(x1, y1, color=['green', 'red'])
    ax1.set_title("Solo Mode Wins and Loses")
    ax1.set_ylabel("Cantidad")

    ax2.pie(y1, explode=(0.1,0), labels=x1, colors=['lightgreen', 'lightcoral'],
            autopct='%1.1f%%', shadow=True, startangle=140)
    ax2.axis('equal')
    ax2.set_title("Solo Mode Wins and Loses")

    ax3.bar(x2, y2, color=['green', 'red'])
    ax3.set_title("Two Players Mode Wins and Loses")
    ax3.set_ylabel("Cantidad")

    ax4.pie(y2, explode=(0.1,0), labels=x2, colors=['lightgreen', 'lightcoral'],
            autopct='%1.1f%%', shadow=True, startangle=140)
    ax4.axis('equal')
    ax4.set_title("Two Players Mode Wins and Loses")

    fig.tight_layout()

    canvas = FigureCanvasTkAgg(fig, master=screen)
    canvas.draw()

    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    toolbar = NavigationToolbar2Tk(canvas, screen)
    toolbar.update()
    toolbar.pack(side=tk.BOTTOM, fill=tk.X)

    median_attempts_frame = tk.Frame(screen, bd=2, relief="groove", bg="lightyellow")
    median_attempts_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
    
    label_median1 = tk.Label(
        median_attempts_frame,
        text=f"Median of attempts Solo Mode: {attempts_mean1}",
        font=("Arial", 16, "bold"),
        bg="lightyellow",
        fg="darkblue"
    )

    label_median2 = tk.Label(
        median_attempts_frame,
        text=f"Median of attempts Two Players Mode: {attempts_mean2}",
        font=("Arial", 16, "bold"),
        bg="lightyellow",
        fg="darkblue"
    )
    label_median1.pack(pady=15)
    label_median2.pack(pady=15)

    window.mainloop()

    





    
    
    

    
